"""Build the DCF workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from dcf_factory.formatting import build_styles, register_styles
from dcf_factory.named_ranges import add_named_range


YEAR_COUNT = 10
YEAR_INDEX_NAMES = [f"year_index_{i}" for i in range(1, YEAR_COUNT + 1)]
YEAR_LABEL_NAMES = [f"year_label_{i}" for i in range(1, YEAR_COUNT + 1)]
YEAR_INDEX_VALUES = [
    "one",
    "two",
    "three",
    "four",
    "five",
    "six",
    "seven",
    "eight",
    "nine",
    "ten",
]

ASSUMPTIONS = [
    ("Period_Start_Year", 2025, "First projection year"),
    ("Projection_Years", 10, "Number of forecast years"),
    ("Discount_Rate", 0.1, "Flat annual discount rate"),
    ("Tax_Rate", 0.25, "Effective tax rate"),
    ("Revenue_Base", 1000000, "Revenue in the most recent historical year"),
    ("Revenue_Growth_Rate", 0.05, "Annual revenue growth rate"),
    ("EBIT_Margin", 0.25, "EBIT as % of revenue"),
    ("Depreciation_Pct_Revenue", 0.03, "Depreciation as % of revenue"),
    ("Capex_Pct_Revenue", 0.04, "Capex as % of revenue"),
    ("NWC_Pct_Revenue", 0.02, "Net working capital as % of revenue"),
    ("Terminal_Method", "Perpetuity Growth", "Terminal value method"),
    ("Terminal_Growth_Rate", 0.03, "Terminal growth rate"),
    ("Net_Debt", 200000, "Net debt"),
    ("Units_Outstanding", 100000, "Units outstanding"),
    ("zero", 0, "Numeric constant"),
    ("one", 1, "Numeric constant"),
    ("two", 2, "Numeric constant"),
    ("three", 3, "Numeric constant"),
    ("four", 4, "Numeric constant"),
    ("five", 5, "Numeric constant"),
    ("six", 6, "Numeric constant"),
    ("seven", 7, "Numeric constant"),
    ("eight", 8, "Numeric constant"),
    ("nine", 9, "Numeric constant"),
    ("ten", 10, "Numeric constant"),
]


def build_dcf(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    styles = build_styles()
    register_styles(workbook, styles)

    inputs_sheet = workbook.active
    inputs_sheet.title = "Inputs"
    operating_sheet = workbook.create_sheet("Operating_Model")
    valuation_sheet = workbook.create_sheet("Valuation")
    outputs_sheet = workbook.create_sheet("Outputs")
    notes_sheet = workbook.create_sheet("Notes")

    _build_inputs(inputs_sheet, workbook, styles)
    _build_operating_model(operating_sheet, styles)
    _build_valuation(valuation_sheet, styles)
    _build_outputs(outputs_sheet, styles)
    _build_notes(notes_sheet, styles)

    workbook.save(output_path)


def _build_inputs(sheet, workbook, styles) -> None:
    sheet["A1"] = "DCF Assumptions"
    sheet["A1"].style = styles["label"]

    sheet["A3"] = "Assumption"
    sheet["B3"] = "Value"
    sheet["C3"] = "Notes"
    for cell in ("A3", "B3", "C3"):
        sheet[cell].style = styles["header"]

    percent_assumptions = {
        "Discount_Rate",
        "Tax_Rate",
        "Revenue_Growth_Rate",
        "EBIT_Margin",
        "Depreciation_Pct_Revenue",
        "Capex_Pct_Revenue",
        "NWC_Pct_Revenue",
        "Terminal_Growth_Rate",
    }

    row = 4
    for name, value, note in ASSUMPTIONS:
        sheet[f"A{row}"] = name
        sheet[f"A{row}"].style = styles["label"]
        sheet[f"B{row}"] = value
        sheet[f"C{row}"] = note
        add_named_range(workbook, name, sheet.title, f"B{row}")
        if name in percent_assumptions:
            sheet[f"B{row}"].style = styles["percent"]
        else:
            sheet[f"B{row}"].style = styles["number"]
        row += 1

    sheet["A23"] = "Year Indices"
    sheet["A23"].style = styles["label"]
    for col, name in enumerate(YEAR_INDEX_NAMES, start=2):
        cell = sheet.cell(row=24, column=col)
        cell.value = f"={YEAR_INDEX_VALUES[col - 2]}"
        add_named_range(workbook, name, sheet.title, cell.coordinate)
        cell.style = styles["number"]

    sheet["A26"] = "Year Labels"
    sheet["A26"].style = styles["label"]
    for col, name in enumerate(YEAR_LABEL_NAMES, start=2):
        cell = sheet.cell(row=27, column=col)
        year_index_name = YEAR_INDEX_NAMES[col - 2]
        cell.value = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"Period_Start_Year+{year_index_name}-one,\"\")"
        )
        add_named_range(workbook, name, sheet.title, cell.coordinate)
        cell.style = styles["number"]

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 40


def _build_operating_model(sheet, styles) -> None:
    sheet["A1"] = "Operating Model"
    sheet["A1"].style = styles["label"]

    sheet["A2"] = "Metric"
    sheet["A2"].style = styles["header"]
    for col, label in enumerate(YEAR_LABEL_NAMES, start=2):
        cell = sheet.cell(row=2, column=col)
        cell.value = f"={label}"
        cell.style = styles["header"]

    metrics = [
        "Revenue",
        "EBIT",
        "NOPAT",
        "Depreciation",
        "Capex",
        "NWC",
        "Change in NWC",
        "Free Cash Flow",
    ]

    for row_offset, metric in enumerate(metrics, start=3):
        sheet[f"A{row_offset}"] = metric
        sheet[f"A{row_offset}"].style = styles["label"]

    for col in range(2, YEAR_COUNT + 2):
        year_col = chr(64 + col)
        year_index_name = YEAR_INDEX_NAMES[col - 2]
        revenue_cell = f"{year_col}3"
        if col == 2:
            revenue_formula = "Revenue_Base*(one+Revenue_Growth_Rate)"
        else:
            prev_col = chr(63 + col)
            revenue_formula = f"{prev_col}3*(one+Revenue_Growth_Rate)"
        sheet[revenue_cell] = (
            f"=IF({year_index_name}<=Projection_Years,{revenue_formula},\"\")"
        )
        sheet[revenue_cell].style = styles["currency"]

        sheet[f"{year_col}4"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}3*EBIT_Margin,\"\")"
        )
        sheet[f"{year_col}4"].style = styles["currency"]

        sheet[f"{year_col}5"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}4*(one-Tax_Rate),\"\")"
        )
        sheet[f"{year_col}5"].style = styles["currency"]

        sheet[f"{year_col}6"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}3*Depreciation_Pct_Revenue,\"\")"
        )
        sheet[f"{year_col}6"].style = styles["currency"]

        sheet[f"{year_col}7"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}3*Capex_Pct_Revenue,\"\")"
        )
        sheet[f"{year_col}7"].style = styles["currency"]

        sheet[f"{year_col}8"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}3*NWC_Pct_Revenue,\"\")"
        )
        sheet[f"{year_col}8"].style = styles["currency"]

        if col == 2:
            nwc_formula = f"{year_col}8-(Revenue_Base*NWC_Pct_Revenue)"
        else:
            prev_col = chr(63 + col)
            nwc_formula = f"{year_col}8-{prev_col}8"
        sheet[f"{year_col}9"] = (
            f"=IF({year_index_name}<=Projection_Years,{nwc_formula},\"\")"
        )
        sheet[f"{year_col}9"].style = styles["currency"]

        sheet[f"{year_col}10"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}5+{year_col}6-{year_col}7-{year_col}9,\"\")"
        )
        sheet[f"{year_col}10"].style = styles["currency"]

    sheet.column_dimensions["A"].width = 28


def _build_valuation(sheet, styles) -> None:
    sheet["A1"] = "Valuation"
    sheet["A1"].style = styles["label"]

    sheet["A2"] = "Metric"
    sheet["A2"].style = styles["header"]

    for col, label in enumerate(YEAR_LABEL_NAMES, start=2):
        cell = sheet.cell(row=2, column=col)
        cell.value = f"={label}"
        cell.style = styles["header"]

    metrics = [
        "Free Cash Flow",
        "Discount Factor",
        "PV of FCF",
        "",
        "Terminal Value",
        "PV of Terminal Value",
        "",
        "PV of Explicit FCF",
        "PV of Terminal Value",
        "Enterprise Value",
        "Net Debt",
        "Equity Value",
        "Units Outstanding",
        "Value Per Unit",
    ]

    for idx, metric in enumerate(metrics, start=3):
        if metric:
            sheet[f"A{idx}"] = metric
            sheet[f"A{idx}"].style = styles["label"]

    for col in range(2, YEAR_COUNT + 2):
        year_col = chr(64 + col)
        year_index_name = YEAR_INDEX_NAMES[col - 2]
        sheet[f"{year_col}3"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"Operating_Model!{year_col}10,\"\")"
        )
        sheet[f"{year_col}3"].style = styles["currency"]

        sheet[f"{year_col}4"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"one/(one+Discount_Rate)^{year_index_name},\"\")"
        )
        sheet[f"{year_col}4"].style = styles["number"]

        sheet[f"{year_col}5"] = (
            f"=IF({year_index_name}<=Projection_Years,"
            f"{year_col}3*{year_col}4,\"\")"
        )
        sheet[f"{year_col}5"].style = styles["currency"]

    last_year_col = chr(64 + YEAR_COUNT + 1)
    sheet[f"{last_year_col}7"] = (
        "=IF(Projection_Years>zero,"
        "INDEX(B3:K3,Projection_Years)*(one+Terminal_Growth_Rate)/"
        "(Discount_Rate-Terminal_Growth_Rate),\"\")"
    )
    sheet[f"{last_year_col}7"].style = styles["currency"]

    sheet[f"{last_year_col}8"] = (
        "=IF(Projection_Years>zero,"
        f"{last_year_col}7*INDEX(B4:K4,Projection_Years),\"\")"
    )
    sheet[f"{last_year_col}8"].style = styles["currency"]

    sheet["B10"] = "=IF(Projection_Years>zero,SUM(B5:K5),\"\")"
    sheet["B10"].style = styles["currency"]

    sheet["B11"] = f"=IF(Projection_Years>zero,{last_year_col}8,\"\")"
    sheet["B11"].style = styles["currency"]

    sheet["B12"] = "=IF(Projection_Years>zero,B10+B11,\"\")"
    sheet["B12"].style = styles["currency"]

    sheet["B13"] = "=Net_Debt"
    sheet["B13"].style = styles["currency"]

    sheet["B14"] = "=B12-B13"
    sheet["B14"].style = styles["currency"]

    sheet["B15"] = "=Units_Outstanding"
    sheet["B15"].style = styles["number"]

    sheet["B16"] = "=IF(Units_Outstanding>zero,B14/B15,\"\")"
    sheet["B16"].style = styles["currency"]

    sheet.column_dimensions["A"].width = 28


def _build_outputs(sheet, styles) -> None:
    sheet["A1"] = "Outputs"
    sheet["A1"].style = styles["label"]

    outputs = [
        ("PV of Explicit FCF", "=Valuation!B10", "currency"),
        ("PV of Terminal Value", "=Valuation!B11", "currency"),
        ("Enterprise Value", "=Valuation!B12", "currency"),
        ("Equity Value", "=Valuation!B14", "currency"),
        ("Value Per Unit", "=Valuation!B16", "currency"),
    ]

    sheet["A3"] = "Metric"
    sheet["B3"] = "Value"
    sheet["A3"].style = styles["header"]
    sheet["B3"].style = styles["header"]

    row = 4
    for label, formula, style_key in outputs:
        sheet[f"A{row}"] = label
        sheet[f"A{row}"].style = styles["label"]
        sheet[f"B{row}"] = formula
        sheet[f"B{row}"].style = styles[style_key]
        row += 1

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 18


def _build_notes(sheet, styles) -> None:
    sheet["A1"] = "Notes"
    sheet["A1"].style = styles["label"]

    lines = [
        "Purpose",
        "This template is designed for educational and illustrative valuation mechanics.",
        "It is intended to help users understand how inputs flow through a discounted",
        "cash flow model in a transparent, single-scenario structure.",
        "",
        "Important Disclosures",
        "- This spreadsheet is for educational use only and does not constitute investment advice.",
        "- It does not guarantee accuracy, completeness, or suitability for any purpose.",
        "- All outputs are mechanically derived from user inputs and do not represent forecasts.",
        "- Users are responsible for validating assumptions and results.",
        "",
        "Definitions",
        "- NOPAT: Net Operating Profit After Tax; EBIT × (1 − Tax_Rate).",
        "- FCF: Free Cash Flow to the Firm; NOPAT + Depreciation − Capex − Change in NWC.",
        "- Terminal Value: Value of cash flows beyond the explicit forecast period",
        "  using a single perpetuity growth method.",
    ]

    for row, line in enumerate(lines, start=3):
        sheet[f"A{row}"] = line
